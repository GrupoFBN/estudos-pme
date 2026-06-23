import openpyxl
import json
import re
import os

def clean_value(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == '' or s == 'None':
        return None
    # Remove currency formatting R$ and non-breaking space
    s_clean = s.replace('R$', '').replace('\xa0', '').replace(' ', '')
    # Handle Brazilian number format: 50.165,81 -> 50165.81
    if re.match(r'^\d{1,3}(\.\d{3})+,\d{2}$', s_clean):
        s_clean = s_clean.replace('.', '').replace(',', '.')
    elif ',' in s_clean and '.' in s_clean:
        s_clean = s_clean.replace('.', '').replace(',', '.')
    elif ',' in s_clean:
        s_clean = s_clean.replace(',', '.')
    try:
        return float(s_clean)
    except:
        return v.strip() if isinstance(v, str) else str(v).strip()

def get_row_values(row):
    return [(j, c) for j, c in enumerate(row) if c is not None and str(c).strip() not in ('', 'None')]

def parse_rows(rows, sheet_name):
    result = {'company': sheet_name.strip(), 'sections': []}
    
    current_section = None
    current_plan_names = None
    current_plan_col_start = None
    current_rows = []
    current_acomodacao = None
    current_abrangencia = None
    current_reembolso = None

    def save_table():
        nonlocal current_plan_names, current_rows, current_plan_col_start
        nonlocal current_acomodacao, current_abrangencia, current_reembolso
        if current_section is not None and current_plan_names and current_rows:
            table = {
                'plans': current_plan_names,
                'age_data': [],
                'total': None,
                'acomodacao': current_acomodacao,
                'abrangencia': current_abrangencia,
                'reembolso': current_reembolso
            }
            for r in current_rows:
                if r.get('is_total'):
                    table['total'] = r['values']
                else:
                    table['age_data'].append(r)
            current_section['tables'].append(table)
        current_plan_names = None
        current_plan_col_start = None
        current_rows = []
        current_acomodacao = None
        current_abrangencia = None
        current_reembolso = None

    for i, row in enumerate(rows):
        cells = get_row_values(row)
        if not cells:
            continue

        first_idx, first_val = cells[0]
        first_str = str(first_val).strip()
        
        # Detect section headers
        section_keywords = ['SEM COPARTICIPA', 'COPARTICIPA', 'LINHA PORTO', 'LINHA TRADICIONAL']
        is_section_header = any(kw in first_str.upper() for kw in section_keywords)
        if not is_section_header and len(cells) > 1:
            second_str = str(cells[1][1]).strip()
            is_section_header = any(kw in second_str.upper() for kw in section_keywords)
            if is_section_header:
                first_str = second_str

        if first_str.startswith('***') or first_str.startswith('*'):
            continue

        # Check for inline sections like in HapVida-Smart
        if not is_section_header:
            inline_sections = [c for j, c in cells if any(kw in str(c).upper() for kw in section_keywords)]
            if inline_sections:
                is_section_header = True
                first_str = str(inline_sections[0]).strip()

        faixa_cells = [(j, c) for j, c in enumerate(row) if c is not None and 'Faixa' in str(c)]
        is_plan_header = bool(faixa_cells)

        total_cells = [(j, c) for j, c in enumerate(row) if c is not None and 'Total' in str(c)]
        is_total_row = bool(total_cells)
        
        is_acomodacao = any('Acomoda' in str(c) for _, c in cells)
        is_abrangencia = any('Abrang' in str(c) for _, c in cells)
        is_reembolso = any('Reembolso' in str(c) for _, c in cells)

        if is_section_header and not is_plan_header:
            save_table()
            current_section = {'name': first_str, 'tables': []}
            result['sections'].append(current_section)

        elif is_plan_header:
            save_table()
            # If the current section is None, we need a default section
            if current_section is None:
                current_section = {'name': 'Planos', 'tables': []}
                result['sections'].append(current_section)

            faixa_col = faixa_cells[0][0]
            plan_names = []
            for j in range(faixa_col + 1, len(row)):
                v = row[j]
                if v is not None and str(v).strip() not in ('', 'None'):
                    if 'Faixa' in str(v):
                        break
                    plan_names.append(str(v).strip())
            current_plan_names = plan_names
            current_plan_col_start = faixa_col + 1

        elif is_total_row and current_plan_names:
            # Handle inline totals like HapVida
            for tc_idx, tc_val in total_cells:
                vals = []
                # Sometimes total is exactly at the start of the columns, sometimes shifted.
                # Find the first column after the 'Total' text
                for j in range(tc_idx + 1, tc_idx + len(current_plan_names) + 5):
                    if j < len(row):
                        v = clean_value(row[j])
                        if v is not None:
                            vals.append(v)
                vals = vals[:len(current_plan_names)]
                if vals:
                    current_rows.append({'label': str(tc_val).strip(), 'values': vals, 'is_total': True})

        elif is_acomodacao and current_plan_names:
            vals = []
            for j, c in cells:
                if 'Acomoda' not in str(c):
                    vals.append(str(c).strip())
            current_acomodacao = vals[:len(current_plan_names)]

        elif is_abrangencia and current_plan_names:
            vals = []
            for j, c in cells:
                if 'Abrang' not in str(c):
                    vals.append(str(c).strip())
            current_abrangencia = vals[:len(current_plan_names)]

        elif is_reembolso and current_plan_names:
            vals = []
            for j, c in cells:
                sv = str(c).strip()
                if 'Reembolso' not in sv and 'consulta' not in sv.lower():
                    cv = clean_value(c)
                    if cv is not None:
                        vals.append(cv)
            if current_reembolso is None:
                current_reembolso = {'PS': vals[:len(current_plan_names)]}
            else:
                current_reembolso['eletiva'] = vals[:len(current_plan_names)]

        elif current_plan_names and current_plan_col_start is not None:
            label_candidates = [(j, c) for j, c in cells if str(c).strip() and not str(c).startswith('R$')]
            if not label_candidates:
                continue
            label_idx, label_val = label_candidates[0]
            label_str = str(label_val).strip()
            
            is_age = re.match(r'^\d', label_str) is not None or 'ou mais' in label_str.lower()
            if is_age:
                vals = []
                for j in range(label_idx + 1, label_idx + len(current_plan_names) + 5):
                    if j < len(row):
                        v = clean_value(row[j])
                        if v is not None:
                            vals.append(v)
                vals = vals[:len(current_plan_names)]
                if vals:
                    current_rows.append({'label': label_str, 'values': vals})

    save_table()
    return result

def extract_idades(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    
    ranges = {
        '0 a 18': 0, '19 a 23': 0, '24 a 28': 0, '29 a 33': 0,
        '34 a 38': 0, '39 a 43': 0, '44 a 48': 0, '49 a 53': 0,
        '54 a 58': 0, '59 ou mais': 0
    }
    
    regimes = {'CLT': 0, 'PJ': 0}
    
    for row in rows:
        if not row or not row[0] or str(row[0]).strip().lower() == 'nome':
            continue
        try:
            idade = int(row[1])
            regime = str(row[2]).strip() if len(row) > 2 and row[2] else 'CLT'
            if regime not in regimes:
                regimes[regime] = 0
            regimes[regime] += 1
            
            if idade <= 18: ranges['0 a 18'] += 1
            elif idade <= 23: ranges['19 a 23'] += 1
            elif idade <= 28: ranges['24 a 28'] += 1
            elif idade <= 33: ranges['29 a 33'] += 1
            elif idade <= 38: ranges['34 a 38'] += 1
            elif idade <= 43: ranges['39 a 43'] += 1
            elif idade <= 48: ranges['44 a 48'] += 1
            elif idade <= 53: ranges['49 a 53'] += 1
            elif idade <= 58: ranges['54 a 58'] += 1
            else: ranges['59 ou mais'] += 1
        except:
            continue

    return {
        'faixas': ranges,
        'regimes': regimes,
        'total_vidas': sum(ranges.values())
    }

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(base_dir, 'ESTUDOS CONVÊNIOS MÉDICOS - XPANSE MATRIZ E FILIAIS.xlsx')
    idades_path = os.path.join(base_dir, 'Idades.xlsx')
    
    idades_data = extract_idades(idades_path)
    print(f"Total de vidas extraídas: {idades_data['total_vidas']}")
    print(f"Regimes: {idades_data['regimes']}")
    
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    all_data = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        # Check if there is a split column (side-by-side tables)
        split_col = None
        for r in rows:
            faixa_indices = [j for j, val in enumerate(r) if val is not None and 'Faixa' in str(val)]
            if len(faixa_indices) >= 2:
                split_col = faixa_indices[1]
                break
        
        if split_col is not None:
            left_rows = [r[:split_col] for r in rows]
            right_rows = [r[split_col:] for r in rows]
            
            left_data = parse_rows(left_rows, sheet_name)
            right_data = parse_rows(right_rows, sheet_name)
            
            company_data = {
                'company': sheet_name.strip(),
                'sections': left_data['sections'] + right_data['sections']
            }
        else:
            company_data = parse_rows(rows, sheet_name)
            
        # Clean up empty sections
        company_data['sections'] = [s for s in company_data['sections'] if s['tables']]
        if company_data['sections']:
            all_data.append(company_data)
            print(f"Company: {company_data['company']} com {len(company_data['sections'])} seções válidas")
            for sec in company_data['sections']:
                print(f"  Section: {sec['name']}")
                for tbl in sec['tables']:
                    print(f"    Plans ({len(tbl['plans'])}): {tbl['plans']}")
                    print(f"    Total: {tbl['total']}")

    out_json = {
        'empresa': 'XPANSE',
        'vidas': idades_data,
        'planos': all_data
    }
    
    out_path = os.path.join(base_dir, 'xpanse_data.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(out_json, f, ensure_ascii=False, indent=2)
    print(f"\nData saved to {out_path}")

if __name__ == '__main__':
    main()
