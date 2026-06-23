"""
FRI Data Extraction - FINAL VERSION
Uses word-level coordinates from pdfplumber to correctly separate:
  - Name column (x ~ 15)
  - Carteirinha column (x ~ 73)
  - CPF column (x ~ 158)
  - Plan code column (x ~ 211)
  - Titularidade + Date + BRL values (x ~ 287+)

Outputs: fri_data.json
"""
import openpyxl
import pdfplumber
import json
import re
import os

# ── HELPERS ─────────────────────────────────────────────────────────────────────

def clean_brl(v):
    if v is None:
        return None
    s = str(v).strip().replace('R$', '').replace('\xa0', '').replace(' ', '')
    if not s or s in ('None', '-', '0,00', '0.00'):
        return None
    if re.match(r'^\d{1,3}(\.\d{3})+,\d{2}$', s):
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except:
        return None


def parse_brl_from_text(text):
    matches = re.findall(r'[\d]+(?:\.\d{3})*,\d{2}', text)
    if matches:
        return clean_brl(matches[-1])
    return None


# ── 1. EXCEL PLANS ──────────────────────────────────────────────────────────────

def extract_excel_plans(excel_path):
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    insurers = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        insurer = {
            'name': sheet_name.strip(),
            'section': None,
            'plans': [],
            'age_data': [],
            'total': None,
            'extra_features': [],
            'notes': []
        }
        plan_names = []
        plan_col_start = None

        for row in rows:
            cells = [(j, v) for j, v in enumerate(row)
                     if v is not None and str(v).strip() and str(v) != 'None']
            if not cells:
                continue

            first_idx, first_val = cells[0]
            first_str = str(first_val).strip()

            upper = first_str.upper()
            if any(kw in upper for kw in ['SEM COPAR', 'COPARTICI']):
                insurer['section'] = 'Sem Coparticipação' if 'SEM COPAR' in upper else 'Coparticipação Parcial'
                continue

            faixa_cells = [(j, c) for j, c in enumerate(row)
                           if c is not None and 'Faixa' in str(c)]
            if faixa_cells:
                faixa_col = faixa_cells[0][0]
                plan_names = []
                for j in range(faixa_col + 1, len(row)):
                    v = row[j]
                    if v is not None and str(v).strip() not in ('', 'None'):
                        plan_names.append(str(v).strip())
                insurer['plans'] = plan_names
                plan_col_start = faixa_col + 1
                continue

            if 'Total' in first_str and plan_names:
                vals = []
                for j in range(first_idx + 1, first_idx + len(plan_names) + 5):
                    if j < len(row):
                        v = clean_brl(row[j])
                        if v is not None:
                            vals.append(v)
                insurer['total'] = vals[:len(plan_names)]
                continue

            if plan_names and plan_col_start is not None:
                is_age = re.match(r'^\d', first_str) or 'ou mais' in first_str.lower()
                if is_age:
                    vals = []
                    for j in range(first_idx + 1, first_idx + len(plan_names) + 5):
                        if j < len(row):
                            v = clean_brl(row[j])
                            if v is not None:
                                vals.append(v)
                    vals = vals[:len(plan_names)]
                    if vals:
                        insurer['age_data'].append({'label': first_str, 'values': vals})
                    continue

                # Extra features and notes extraction below total
                first_str_lower = first_str.lower()
                is_feature = False
                for kw in ['acomoda', 'abrang', 'reembolso', 'reemb']:
                    if kw in first_str_lower:
                        if len(first_str) < 35 and not first_str.startswith('*'):
                            is_feature = True
                            break

                if is_feature:
                    vals = []
                    for j in range(plan_col_start, plan_col_start + len(plan_names)):
                        if j < len(row):
                            v = row[j]
                            vals.append(str(v).strip() if v is not None else '—')
                        else:
                            vals.append('—')
                    # Only add if it actually contains data
                    if any(v != '—' and v.strip() for v in vals):
                        insurer['extra_features'].append({
                            'label': first_str,
                            'values': vals
                        })
                    else:
                        insurer['notes'].append(first_str)
                else:
                    non_empty = [str(x).strip() for x in row if x is not None and str(x).strip() and str(x) != 'None']
                    if len(non_empty) == 1:
                        note = non_empty[0]
                        if note.replace('*', '').strip():
                            insurer['notes'].append(note)

        insurers.append(insurer)
    return insurers


# ── 2. RLA PDFs - using word coordinates ─────────────────────────────────────────

# Column X boundaries (based on analysis)
COL_NAME_MAX   = 73    # Name column: x0 < 73
COL_CART_MIN   = 73    # Carteirinha: 73 <= x0 < 155
COL_CART_MAX   = 155
COL_CPF_MIN    = 155   # CPF: 155 <= x0 < 210
COL_CPF_MAX    = 210
COL_PLAN_MIN   = 210   # Plan code: 210 <= x0 < 285
COL_PLAN_MAX   = 285
COL_DATA_MIN   = 285   # Titularidade/Date/Values: x0 >= 285

BRL_RE = re.compile(r'BRL([\d]+(?:\.[\d]{3})*,\d{2})')


def extract_rla_coordinated(pdf_path):
    """Parse RLA PDF using word coordinates to correctly assign columns."""

    result = {
        'source_file': os.path.basename(pdf_path),
        'entity_name': '',
        'subestipulante': '',
        'competencia': '',
        'invoice_total': None,
        'premio_real': None,
        'iof': None,
        'titulares': 0,
        'dependentes': 0,
        'total_vidas': 0,
        'persons': []
    }

    with pdfplumber.open(pdf_path) as pdf:
        all_pages_text = '\n'.join(p.extract_text() or '' for p in pdf.pages)
        all_words_by_page = [p.extract_words() for p in pdf.pages]

    # ── Summary from plain text ──
    for ln in all_pages_text.split('\n'):
        ln = ln.strip()
        if 'Estipulante:' in ln:
            m = re.search(r'Estipulante:\s*\d+\s+(.+?)\s+Fatura:', ln)
            if m:
                result['entity_name'] = m.group(1).strip()
        if 'Subestipulante:' in ln:
            m = re.search(r'Subestipulante:\s*\d+\s+(.+)', ln)
            if m:
                result['subestipulante'] = m.group(1).strip()
        if re.search(r'refer[eê]ncia', ln, re.I):
            m = re.search(r'([A-Za-z\u00e0-\u00ff]+/\d{4})', ln)
            if m:
                result['competencia'] = m.group(1)
        if re.search(r'mio Real', ln, re.I):
            v = parse_brl_from_text(ln)
            if v:
                result['premio_real'] = v
        if 'I.O.F' in ln:
            v = parse_brl_from_text(ln)
            if v:
                result['iof'] = v
        if re.match(r'Total\s+[\d\.,]+', ln) and result['invoice_total'] is None:
            v = parse_brl_from_text(ln)
            if v and v > 100:
                result['invoice_total'] = v
        m = re.search(r'Titulares\s*\.+\s*(\d+)', ln)
        if m:
            result['titulares'] = int(m.group(1))
        m = re.search(r'Dependentes\s*\.+\s*(\d+)', ln)
        if m:
            result['dependentes'] = int(m.group(1))
        m = re.match(r'^Total\s+(\d+)$', ln)
        if m:
            result['total_vidas'] = int(m.group(1))

    # ── Per-person data using word coordinates ──
    # Group words by page, then by approximate Y-row (within 2px tolerance)
    # Key: the main data line for each person has: carteirinha, cpf, titularidade, date, BRL values
    # That line has x0 ~ 73 (carteirinha start)

    persons = []
    
    # Track "current person block" across rows
    # A person block starts when we see a carteirinha (16 digits, x0~73)
    # Preceding rows (same or nearby y) give us name and plan codes

    for page_words in all_words_by_page:
        # Sort by y then x
        sorted_words = sorted(page_words, key=lambda w: (round(w['top'], 0), w['x0']))
        
        # Group into rows (same top within 2px)
        rows_dict = {}
        for w in sorted_words:
            y_key = round(w['top'] / 2) * 2  # round to nearest 2
            if y_key not in rows_dict:
                rows_dict[y_key] = []
            rows_dict[y_key].append(w)
        
        rows = [(y, words) for y, words in sorted(rows_dict.items())]
        
        # Find person rows (those containing carteirinha + cpf)
        person_row_indices = []
        for ri, (y, rwords) in enumerate(rows):
            texts = [w['text'] for w in rwords]
            full = ' '.join(texts)
            # A data row has carteirinha (10-16 digit number) and CPF (11 digit) and date pattern
            if (re.search(r'\d{10,16}', full) and 
                re.search(r'\d{11}', full) and 
                re.search(r'(Titular|Dependente)', full) and
                re.search(r'\d{2}/\d{2}/\d{4}', full)):
                person_row_indices.append(ri)
        
        # For each person row, look back until the previous person row for name/plan context
        prev_person_ri = -1
        for ri in person_row_indices:
            y, rwords = rows[ri]
            full_row = ' '.join(w['text'] for w in rwords)
            
            # Extract from data row
            cart_m = re.search(r'(\d{9,16})', full_row)
            cpf_m = re.search(r'(\d{11})', full_row)
            tit_m = re.search(r'(Titular|Dependente)', full_row)
            date_m = re.search(r'(\d{2}/\d{2}/\d{4})', full_row)
            brl_vals = BRL_RE.findall(full_row)
            
            if not (cart_m and cpf_m and tit_m and date_m):
                prev_person_ri = ri
                continue
            
            carteirinha = cart_m.group(1)
            cpf = cpf_m.group(1)[:11]
            titularidade = tit_m.group(1)
            date = date_m.group(1)
            
            # BRL values: premio, pro_rata, copar, iof, valor
            valor = clean_brl(brl_vals[-1]) if brl_vals else None
            premio = clean_brl(brl_vals[0]) if brl_vals else None
            
            # Look back for name and plan code
            # IMPORTANT: only look back to the previous person row to avoid cross-contamination
            name_words = []
            plan_blocks = []
            
            look_back = prev_person_ri + 1  # start from row AFTER previous person
            prev_plan_y = None
            current_plan_block = []
            
            for prev_ri in range(look_back, ri):
                prev_y, pwords = rows[prev_ri]
                
                name_col_words = [pw for pw in pwords if pw['x0'] < COL_NAME_MAX 
                                  and re.match(r'^[A-Z\u00c0-\u00dc]+$', pw['text'])]
                plan_col_words = [pw for pw in pwords if COL_PLAN_MIN <= pw['x0'] < COL_PLAN_MAX]
                
                if name_col_words:
                    for pw in name_col_words:
                        if abs(pw['top'] - y) < 60:
                            name_words.append((pw['top'], pw['text']))
                
                if plan_col_words:
                    # Check if this is a new plan block (starts with 4-digit code)
                    plan_text = ' '.join(pw['text'] for pw in plan_col_words)
                    starts_new = bool(re.match(r'^\d{4}', plan_text))
                    
                    if starts_new and current_plan_block:
                        # Save current block and start new one
                        plan_blocks.append(current_plan_block[:])
                        current_plan_block = [(prev_y, plan_text)]
                    elif starts_new:
                        current_plan_block = [(prev_y, plan_text)]
                    elif current_plan_block:
                        current_plan_block.append((prev_y, plan_text))
            
            if current_plan_block:
                plan_blocks.append(current_plan_block)
            
            # Also check the data row itself for name words
            for pw in rwords:
                if pw['x0'] < COL_NAME_MAX and re.match(r'^[A-Z\u00c0-\u00dc]+$', pw['text']):
                    name_words.append((pw['top'], pw['text']))
            
            # Filter and sort name_words
            name_words = [(y2, t) for y2, t in name_words if abs(y2 - y) < 65]
            name_words.sort(key=lambda x: x[0])
            name = ' '.join(t for _, t in name_words).strip()
            
            # Porto Saude plan code mapping (from RLA document analysis)
            # Supplemental/secondary codes to ignore:
            SECONDARY_CODES = {'9027', '9627', '9027BRONZE', '9627VIDA'}
            
            # Find primary plan from all collected plan blocks
            plan_name = ''
            plan_raw = ''
            
            all_block_texts = []
            for blk in plan_blocks:
                all_block_texts.extend([t for _, t in blk])
            
            # Try each plan block: find first one NOT starting with a secondary code
            for blk in plan_blocks:
                blk_text = ' '.join(t for _, t in blk)
                code_match = re.match(r'(\d{4})(.*)', blk_text)
                if code_match:
                    code = code_match.group(1)
                    rest = code_match.group(2).strip()
                    # Check if it's a secondary code
                    if code in {'9027', '9627'}:
                        continue
                    # This is the primary plan
                    plan_raw = blk_text
                    plan_name = rest
                    # Clean up
                    plan_name = re.sub(r'\s*/\s*$', '', plan_name).strip()
                    plan_name = re.sub(r'\s+', ' ', plan_name).strip()
                    plan_name = re.sub(r'\s+(BRASIL\s+E|BCO\s+AP|RD\s+PORTO|BRANCO\s+AP).*$', '', plan_name).strip()
                    plan_name = re.sub(r'\s+MEDISERVICE.*$', '', plan_name).strip()
                    break
            
            # Fallback: check if plan name is embedded in CPF string (sometimes glued)
            if not plan_name:
                cpf_raw_full = cpf_m.group(1) if cpf_m else ''
                if len(cpf_raw_full) > 11:
                    extra = cpf_raw_full[11:]
                    m3 = re.match(r'\d{4}(.+)', extra)
                    if m3:
                        plan_name = m3.group(1).strip()
                        plan_name = re.sub(r'\s+(BRASIL\s+E|BCO\s+AP|RD\s+PORTO).*$', '', plan_name).strip()
            
            persons.append({
                'name': name[:60].strip(),
                'carteirinha': carteirinha,
                'cpf': cpf,
                'titularidade': titularidade,
                'inicio_vigencia': date,
                'plano_raw': plan_raw[:80],
                'plano': plan_name[:60],
                'valor': valor,
                'premio': premio,
            })
            prev_person_ri = ri  # update boundary for next person

    result['persons'] = persons
    return result


# ── MAIN ────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    base_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(base_dir, '..', 'temp_excel', 'fri_study.xlsx')
    rla1_path  = os.path.join(base_dir, 'RLA_22_06_2026.pdf')
    rla2_path  = os.path.join(base_dir, 'RLA_22_06_2026 (1).pdf')
    rla3_path  = os.path.join(base_dir, 'RLA_23_06_2026 (1).pdf')

    print("Extracting Excel plans...")
    plans = extract_excel_plans(excel_path)
    for p in plans:
        print(f"  {p['name']}: {len(p['plans'])} planos | {len(p['age_data'])} faixas | {p['section']}")

    # ── RLA 1: FRI (Empresa Principal / Estipulante) ──
    print("\nExtracting RLA 1 (FRI - Empresa Principal)...")
    rla1 = extract_rla_coordinated(rla1_path)
    print(f"  Empresa: {rla1['entity_name']}")
    print(f"  Subestipulante: {rla1['subestipulante']}")
    print(f"  Competencia: {rla1['competencia']}")
    print(f"  Total: R$ {rla1['invoice_total']}")
    print(f"  Vidas: {rla1['total_vidas']} ({rla1['titulares']}T + {rla1['dependentes']}D)")
    print(f"  Persons found: {len(rla1['persons'])}")
    for p in rla1['persons'][:10]:
        print(f"    [{p['titularidade'][:1]}] {p['name'][:28]:28s} | {p['plano'][:22]:22s} | R${p['valor']}")

    # ── RLA 2: KCB (Subestipulante) ──
    print("\nExtracting RLA 2 (KCB Subestipulante)...")
    rla2 = extract_rla_coordinated(rla2_path)
    print(f"  Subestipulante: {rla2['subestipulante']}")
    print(f"  Total: R$ {rla2['invoice_total']}")
    print(f"  Vidas: {rla2['total_vidas']}")
    print(f"  Persons found: {len(rla2['persons'])}")
    for p in rla2['persons']:
        print(f"    [{p['titularidade'][:1]}] {p['name'][:30]:30s} | {p['plano'][:25]:25s} | R${p['valor']}")

    # ── RLA 3: Aura Cosmeticos (Subestipulante) ──
    print("\nExtracting RLA 3 (Aura Cosmeticos Subestipulante)...")
    rla3 = extract_rla_coordinated(rla3_path)
    print(f"  Subestipulante: {rla3['subestipulante']}")
    print(f"  Total: R$ {rla3['invoice_total']}")
    print(f"  Vidas: {rla3['total_vidas']}")
    print(f"  Persons found: {len(rla3['persons'])}")
    for p in rla3['persons']:
        print(f"    [{p['titularidade'][:1]}] {p['name'][:30]:30s} | {p['plano'][:25]:25s} | R${p['valor']}")

    all_rlas = [rla1, rla2, rla3]
    total_combined = sum((r['invoice_total'] or 0) for r in all_rlas)
    total_vidas = sum(r['total_vidas'] for r in all_rlas)

    output = {
        'case': 'FRI',
        'empresa_name': 'FRI Comercio de Cosmeticos e Perfumaria',
        'competencia': rla1['competencia'] or rla2['competencia'] or rla3['competencia'],
        'seguradora_atual': 'Porto Saude S.A.',
        'fatura_total_atual': total_combined,
        'total_vidas_atual': total_vidas,
        'rla_empresa': rla1,
        'rla_subestipulante': rla2,
        'rla_subs': [
            {'label': 'FRI Comercio de Cosmeticos e Perfumaria', 'cnpj': '23.127.599/0001-30', 'data': rla1},
            {'label': 'KCB Serviços de Apoio', 'cnpj': '', 'data': rla2},
            {'label': 'Aura Cosmeticos e Perfumaria LTDA', 'cnpj': '49.532.571/0001-66', 'data': rla3},
        ],
        'plans_excel': plans
    }

    out_path = os.path.join(base_dir, 'fri_data.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"\nSalvo em: {out_path}")
    print(f"Fatura total combinada: R$ {total_combined:,.2f}")
    print(f"Total vidas: {total_vidas}")
