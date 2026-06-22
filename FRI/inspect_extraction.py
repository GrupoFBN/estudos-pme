import openpyxl
import re
import os

def clean_brl(v):
    if v is None:
        return None
    s = str(v).strip().replace('R$', '').replace('\xa0', '').replace(' ', '')
    if not s or s in ('None', '-', '0,00', '0.00'):
        return None
    if re.match(r'^\d{1,3}(\.\d{3})+,\d{2}$', s):
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except:
        return None

excel_path = r"c:\Users\Danilo\OneDrive - FBN RE Administradora e Corretora de Seguros Ltda\Área de Trabalho\Estudos PME\temp_excel\fri_study.xlsx"
wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))

    insurer = {
        'name': name.strip(),
        'section': None,
        'plans': [],
        'age_data': [],
        'total': None,
        'extra_features': [],
        'notes': []
    }
    plan_names = []
    plan_col_start = None

    for row in rows:
        cells = [(j, v) for j, v in enumerate(row)
                 if v is not None and str(v).strip() and str(v) != 'None']
        if not cells:
            continue

        first_idx, first_val = cells[0]
        first_str = str(first_val).strip()

        upper = first_str.upper()
        if any(kw in upper for kw in ['SEM COPAR', 'COPARTICI']):
            insurer['section'] = 'Sem Coparticipação' if 'SEM COPAR' in upper else 'Coparticipação Parcial'
            continue

        faixa_cells = [(j, c) for j, c in enumerate(row)
                       if c is not None and 'Faixa' in str(c)]
        if faixa_cells:
            faixa_col = faixa_cells[0][0]
            plan_names = []
            for j in range(faixa_col + 1, len(row)):
                v = row[j]
                if v is not None and str(v).strip() not in ('', 'None'):
                    plan_names.append(str(v).strip())
            insurer['plans'] = plan_names
            plan_col_start = faixa_col + 1
            continue

        if 'Total' in first_str and plan_names:
            vals = []
            for j in range(first_idx + 1, first_idx + len(plan_names) + 5):
                if j < len(row):
                    v = clean_brl(row[j])
                    if v is not None:
                        vals.append(v)
            insurer['total'] = vals[:len(plan_names)]
            continue

        if plan_names and plan_col_start is not None:
            is_age = re.match(r'^\d', first_str) or 'ou mais' in first_str.lower()
            if is_age:
                vals = []
                for j in range(first_idx + 1, first_idx + len(plan_names) + 5):
                    if j < len(row):
                        v = clean_brl(row[j])
                        if v is not None:
                            v_clean = v
                        else:
                            # Try to keep original if clean_brl fails, e.g. text
                            v_clean = row[j]
                        if v_clean is not None:
                            vals.append(v_clean)
                vals = vals[:len(plan_names)]
                if vals:
                    insurer['age_data'].append({'label': first_str, 'values': vals})
                continue
            
            # If not age, but we have plan names, check if it's extra features or notes
            first_str_lower = first_str.lower()
            is_feature = False
            for kw in ['acomoda', 'abrang', 'reembolso', 'reemb']:
                if kw in first_str_lower:
                    if len(first_str) < 35 and not first_str.startswith('*'):
                        is_feature = True
                        break
            
            if is_feature:
                vals = []
                for j in range(plan_col_start, plan_col_start + len(plan_names)):
                    if j < len(row):
                        v = row[j]
                        vals.append(str(v).strip() if v is not None else '—')
                    else:
                        vals.append('—')
                # Only add if it actually has values
                if any(v != '—' and v.strip() for v in vals):
                    insurer['extra_features'].append({
                        'label': first_str,
                        'values': vals
                    })
                else:
                    insurer['notes'].append(first_str)
            else:
                non_empty = [str(x).strip() for x in row if x is not None and str(x).strip() and str(x) != 'None']
                if len(non_empty) == 1:
                    note = non_empty[0]
                    if note.replace('*', '').strip():
                        insurer['notes'].append(note)

    print(f"=== {insurer['name']} ===")
    print(f"  Section: {insurer['section']}")
    print(f"  Plans: {insurer['plans']}")
    print(f"  Features:")
    for f in insurer['extra_features']:
        print(f"    {f['label']}: {f['values']}")
    print(f"  Notes:")
    for n in insurer['notes']:
        print(f"    - {n}")
    print()
