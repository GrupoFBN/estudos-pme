import openpyxl
import json
import re

fname = 'ESTUDOS CONVÊNIOS MÉDICOS - GRUPO SALUS.xlsx'
wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)

def clean_value(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == '' or s == 'None':
        return None
    # Remove currency formatting R$ and non-breaking space
    s_clean = s.replace('R$', '').replace('\xa0', '').replace(' ', '')
    # Handle Brazilian number format: 50.165,81 -> 50165.81
    # Check if it matches pattern with dots as thousands separators
    if re.match(r'^\d{1,3}(\.\d{3})+,\d{2}$', s_clean):
        s_clean = s_clean.replace('.', '').replace(',', '.')
    elif ',' in s_clean and '.' in s_clean:
        # Assume comma is decimal and dots are thousands
        s_clean = s_clean.replace('.', '').replace(',', '.')
    elif ',' in s_clean:
        s_clean = s_clean.replace(',', '.')
    try:
        return float(s_clean)
    except:
        return v.strip() if isinstance(v, str) else str(v).strip()


def get_row_values(row):
    """Returns list of (col_index, value) for non-empty cells"""
    return [(j, c) for j, c in enumerate(row) if c is not None and str(c).strip() not in ('', 'None')]


def parse_company(ws, sheet_name):
    rows = list(ws.iter_rows(values_only=True))
    result = {'company': sheet_name.strip(), 'sections': []}
    
    current_section = None
    current_plan_names = None
    current_plan_col_start = None
    current_rows = []
    current_acomodacao = None
    current_abrangencia = None
    current_reembolso = None

    def save_table():
        nonlocal current_plan_names, current_rows, current_plan_col_start
        nonlocal current_acomodacao, current_abrangencia, current_reembolso
        if current_section is not None and current_plan_names and current_rows:
            table = {
                'plans': current_plan_names,
                'age_data': [],
                'total': None,
                'acomodacao': current_acomodacao,
                'abrangencia': current_abrangencia,
                'reembolso': current_reembolso
            }
            for r in current_rows:
                if r.get('is_total'):
                    table['total'] = r['values']
                else:
                    table['age_data'].append(r)
            current_section['tables'].append(table)
        current_plan_names = None
        current_plan_col_start = None
        current_rows = []
        current_acomodacao = None
        current_abrangencia = None
        current_reembolso = None

    for i, row in enumerate(rows):
        cells = get_row_values(row)
        if not cells:
            continue

        first_idx, first_val = cells[0]
        first_str = str(first_val).strip()
        
        # Detect section headers
        section_keywords = ['SEM COPARTICIPA', 'COPARTICIPA', 'LINHA PORTO', 'LINHA TRADICIONAL']
        is_section_header = any(kw in first_str.upper() for kw in section_keywords)
        # Also check second column
        if not is_section_header and len(cells) > 1:
            second_str = str(cells[1][1]).strip()
            is_section_header = any(kw in second_str.upper() for kw in section_keywords)
            if is_section_header:
                first_str = second_str

        # Also ignore lines that start with ***
        if first_str.startswith('***') or first_str.startswith('*'):
            continue

        # Detect plan header row (contains "Faixa")
        faixa_cells = [(j, c) for j, c in enumerate(row) if c is not None and 'Faixa' in str(c)]
        is_plan_header = bool(faixa_cells)

        # Detect total row
        total_cells = [(j, c) for j, c in enumerate(row) if c is not None and 'Total' in str(c)]
        is_total_row = bool(total_cells)
        
        # Detect metadata rows
        is_acomodacao = any('Acomoda' in str(c) for _, c in cells)
        is_abrangencia = any('Abrang' in str(c) for _, c in cells)
        is_reembolso = any('Reembolso' in str(c) for _, c in cells)

        if is_section_header and not is_plan_header:
            save_table()
            if current_section is not None:
                pass  # already in result
            current_section = {'name': first_str, 'tables': []}
            result['sections'].append(current_section)

        elif is_plan_header:
            save_table()
            # Extract plan names from columns after "Faixa Etária"
            faixa_col = faixa_cells[0][0]
            plan_names = []
            for j in range(faixa_col + 1, len(row)):
                v = row[j]
                if v is not None and str(v).strip() not in ('', 'None'):
                    plan_names.append(str(v).strip())
            current_plan_names = plan_names
            current_plan_col_start = faixa_col + 1

        elif is_total_row and current_plan_names:
            for tc_idx, tc_val in total_cells:
                # Take values after "Total" cell
                vals = []
                for j in range(tc_idx + 1, tc_idx + len(current_plan_names) + 5):
                    if j < len(row):
                        v = clean_value(row[j])
                        if v is not None:
                            vals.append(v)
                vals = vals[:len(current_plan_names)]
                if vals:
                    current_rows.append({'label': 'Total 66 Vidas', 'values': vals, 'is_total': True})
                break

        elif is_acomodacao and current_plan_names:
            vals = []
            for j, c in cells:
                if 'Acomoda' not in str(c):
                    vals.append(str(c).strip())
            current_acomodacao = vals[:len(current_plan_names)]

        elif is_abrangencia and current_plan_names:
            vals = []
            for j, c in cells:
                if 'Abrang' not in str(c):
                    vals.append(str(c).strip())
            current_abrangencia = vals[:len(current_plan_names)]

        elif is_reembolso and current_plan_names:
            vals = []
            for j, c in cells:
                sv = str(c).strip()
                if 'Reembolso' not in sv and 'consulta' not in sv.lower():
                    cv = clean_value(c)
                    if cv is not None:
                        vals.append(cv)
            if current_reembolso is None:
                current_reembolso = {'PS': vals[:len(current_plan_names)]}
            else:
                current_reembolso['eletiva'] = vals[:len(current_plan_names)]

        elif current_plan_names and current_plan_col_start is not None:
            # Try to parse as age range data row
            # Look for the first meaningful label
            label_candidates = [(j, c) for j, c in cells if str(c).strip() and not str(c).startswith('R$')]
            if not label_candidates:
                continue
            label_idx, label_val = label_candidates[0]
            label_str = str(label_val).strip()
            
            # Check if it's an age-range-like label
            is_age = re.match(r'^\d', label_str) is not None or 'ou mais' in label_str.lower()
            if is_age:
                # Get values after label
                vals = []
                for j in range(label_idx + 1, label_idx + len(current_plan_names) + 5):
                    if j < len(row):
                        v = clean_value(row[j])
                        if v is not None:
                            vals.append(v)
                vals = vals[:len(current_plan_names)]
                if vals:
                    current_rows.append({'label': label_str, 'values': vals})

    save_table()
    return result


all_data = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    company_data = parse_company(ws, sheet_name)
    all_data.append(company_data)
    print(f"Company: {company_data['company']}")
    for sec in company_data['sections']:
        print(f"  Section: {sec['name']}")
        for tbl in sec['tables']:
            print(f"    Plans ({len(tbl['plans'])}): {tbl['plans']}")
            print(f"    Age rows: {len(tbl['age_data'])}")
            print(f"    Total: {tbl['total']}")
            print(f"    Acomodacao: {tbl['acomodacao']}")

with open('health_data.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)
print("\nData saved to health_data.json")
